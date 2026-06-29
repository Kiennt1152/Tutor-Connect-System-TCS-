"""Generate UC → Entity mapping Excel for TCS (71 UC)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "UC ID",
    "Module",
    "UC Name",
    "Primary Actor",
    "Entity Name",
    "Attribute",
    "State",
]

ROWS = [
    # Authentication & RBAC
    ("UC-01", "Authentication & RBAC", "Login / Logout", "Client, Tutor, Tutor Center, Platform Admin", "users", "user_id, email, phone, password_hash, status, created_at, updated_at, last_login", "status: ACTIVE, SUSPENDED, BANNED"),
    ("UC-02", "Authentication & RBAC", "Register Account", "Client, Tutor, Tutor Center", "users", "user_id, email, phone, password_hash, status, created_at", "status: ACTIVE, SUSPENDED, BANNED"),
    ("", "", "", "", "clients", "client_id, user_id, full_name, phone, date_of_birth, gender, address, location_id, avatar_url, created_at, updated_at", "gender: MALE, FEMALE, OTHER"),
    ("", "", "", "", "tutors", "tutor_id, user_id, full_name, gender, date_of_birth, phone, address, location_id, bio, avatar, experience_years, hourly_rate, rating_avg, verification_status, created_at, updated_at", "gender: MALE, FEMALE, OTHER; verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("", "", "", "", "tutor_centers", "center_id, user_id, company_name, license_no, phone, address, location_id, description, avatar, verification_status, created_at, updated_at", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("UC-03", "Authentication & RBAC", "Reset Password", "Client, Tutor, Tutor Center, Platform Admin", "password_reset_tokens", "token_id, user_id, token, expires_at, used_at, created_at", ""),
    ("", "", "", "", "users", "user_id, email, password_hash", ""),
    ("UC-04", "Authentication & RBAC", "Change Password", "Client, Tutor, Tutor Center, Platform Admin", "users", "user_id, password_hash", ""),
    ("UC-05", "Authentication & RBAC", "Link Parent-Child Account", "Client", "parent_child_links", "link_id, parent_user_id, child_profile_id, status, created_at", "status: ACTIVE, REVOKED"),
    ("UC-06", "Authentication & RBAC", "Manage Child Profiles", "Client", "child_profiles", "child_profile_id, full_name, date_of_birth, gender, grade_id, school_name, notes, created_at, updated_at", "gender: MALE, FEMALE, OTHER"),
    ("", "", "", "", "grades", "grade_id, grade_name", ""),
    ("UC-07", "Authentication & RBAC", "Manage User Accounts", "Platform Admin", "users", "user_id, email, phone, status, created_at, updated_at", "status: ACTIVE, SUSPENDED, BANNED"),
    ("", "", "", "", "user_penalties", "penalty_id, user_id, issued_by, penalty_type, reason, evidence_urls, restriction_details, starts_at, expires_at, status, revoked_at, revoked_reason, created_at", "penalty_type: WARNING, FEATURE_RESTRICTION, TEMPORARY_BAN, PERMANENT_BAN; status: ACTIVE, EXPIRED, REVOKED"),
    ("", "", "", "", "audit_logs", "audit_id, actor_id, action, entity_type, entity_id, old_value, new_value, ip_address, user_agent, created_at", ""),
    # Profile Verification
    ("UC-08", "Profile Verification Workflow", "Manage Personal Profile", "Client, Tutor, Tutor Center, Platform Admin", "clients", "client_id, user_id, full_name, phone, date_of_birth, gender, address, location_id, avatar_url, created_at, updated_at", "gender: MALE, FEMALE, OTHER"),
    ("", "", "", "", "tutors", "tutor_id, user_id, full_name, gender, date_of_birth, phone, address, location_id, bio, avatar, experience_years, hourly_rate, rating_avg, verification_status, created_at, updated_at", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("", "", "", "", "tutor_centers", "center_id, user_id, company_name, license_no, phone, address, location_id, description, avatar, verification_status, created_at, updated_at", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("", "", "", "", "platform_admins", "admin_id, user_id, full_name, created_at, updated_at", ""),
    ("", "", "", "", "media_files", "file_id, user_id, file_name, file_url, mime_type, file_size, created_at", ""),
    ("", "", "", "", "locations", "location_id, google_place_id, address_line, ward_name, district_name, province_id, latitude, longitude", ""),
    ("UC-09", "Profile Verification Workflow", "Manage Tutor Verifications", "Tutor", "verification_requests", "verification_id, user_id, verification_type, admin_notes, status, submitted_at, reviewed_at, created_at, updated_at", "verification_type: TUTOR_PROFILE, TUTOR_CENTER_LICENSE; status: DRAFT, SUBMITTED, UNDER_REVIEW, VERIFIED, REJECTED"),
    ("", "", "", "", "verification_documents", "document_id, verification_id, file_id, document_type", "document_type: ID_CARD, DEGREE, CERTIFICATE, LICENSE"),
    ("", "", "", "", "media_files", "file_id, file_url, mime_type", ""),
    ("", "", "", "", "qualifications", "qualification_id, tutor_id, title, issuer, issue_date", ""),
    ("", "", "", "", "tutor_educations", "education_id, tutor_id, institution, degree, field_of_study, start_year, end_year", ""),
    ("", "", "", "", "tutor_certificates", "certificate_id, tutor_id, name, issuer, issue_date, file_id", ""),
    ("", "", "", "", "tutor_experiences", "experience_id, tutor_id, organization, role, start_date, end_date, description", ""),
    ("UC-10", "Profile Verification Workflow", "Manage Center License", "Tutor Center", "verification_requests", "verification_id, user_id, verification_type, admin_notes, status, submitted_at, reviewed_at", "verification_type: TUTOR_CENTER_LICENSE; status: DRAFT, SUBMITTED, UNDER_REVIEW, VERIFIED, REJECTED"),
    ("", "", "", "", "verification_documents", "document_id, verification_id, file_id, document_type", "document_type: ID_CARD, DEGREE, CERTIFICATE, LICENSE"),
    ("", "", "", "", "tutor_centers", "center_id, user_id, license_no, verification_status", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("UC-11", "Profile Verification Workflow", "Manage Document Verifications", "Platform Admin", "verification_requests", "verification_id, user_id, verification_type, admin_notes, status, submitted_at, reviewed_at, created_at, updated_at", "status: DRAFT, SUBMITTED, UNDER_REVIEW, VERIFIED, REJECTED"),
    ("", "", "", "", "verification_documents", "document_id, verification_id, file_id, document_type", "document_type: ID_CARD, DEGREE, CERTIFICATE, LICENSE"),
    ("", "", "", "", "verification_histories", "verification_history_id, verification_id, old_status, new_status, changed_by_user_id, changed_at", ""),
    ("UC-12", "Profile Verification Workflow", "Manage Teaching Availability", "Tutor", "tutor_availabilities", "availability_id, tutor_id, day_of_week, start_time, end_time, is_recurring, specific_date, google_calendar_event_id, created_at, updated_at", ""),
    ("UC-13", "Profile Verification Workflow", "Sync with Google Calendar", "Tutor", "tutor_availabilities", "availability_id, tutor_id, google_calendar_event_id, day_of_week, start_time, end_time", ""),
    # Class Life-cycle
    ("UC-14", "Class Life-cycle", "Manage Class", "Client, Tutor Center", "tutoring_classes", "class_id, creator_id, category_id, subject_id, grade_id, location_id, title, description, lesson_mode, number_of_sessions, start_date, end_date, tuition_fee, budget, recurring_type, status, created_at, updated_at", "lesson_mode: ONLINE, OFFLINE, HYBRID; recurring_type: ONCE, WEEKLY; status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("UC-15", "Class Life-cycle", "Assign Tutor to Class", "Tutor Center", "class_assignments", "assignment_id, application_id, assigned_date, status", "status: ACTIVE, TERMINATED"),
    ("", "", "", "", "tutor_applications", "application_id, class_id, tutor_id, proposed_rate, cover_letter, status, applied_at, reviewed_at", "status: SUBMITTED, UNDER_REVIEW, ACCEPTED, REJECTED, WITHDRAWN"),
    ("UC-16", "Class Life-cycle", "Apply for Class", "Tutor", "tutor_applications", "application_id, class_id, tutor_id, proposed_rate, cover_letter, status, applied_at, reviewed_at", "status: SUBMITTED, UNDER_REVIEW, ACCEPTED, REJECTED, WITHDRAWN"),
    ("UC-17", "Class Life-cycle", "View Tutor Application for Class", "Client, Tutor Center", "tutor_applications", "application_id, class_id, tutor_id, proposed_rate, cover_letter, status, applied_at, reviewed_at", "status: SUBMITTED, UNDER_REVIEW, ACCEPTED, REJECTED, WITHDRAWN"),
    ("", "", "", "", "application_status_histories", "application_history_id, application_id, old_status, new_status, changed_by_user_id, changed_at", ""),
    ("UC-18", "Class Life-cycle", "Manage Lesson Schedule", "Tutor Center", "schedule_slots", "slot_id, class_id, day_of_week, start_time, end_time", ""),
    ("", "", "", "", "lessons", "lesson_id, class_id, slot_id, sequence_no, tutor_check_in_at, tutor_check_out_at, client_confirm_at, attendance_status, created_at", "attendance_status: PENDING, COMPLETED, ABSENT, DISPUTED"),
    ("UC-19", "Class Life-cycle", "Import Student via Excel", "Tutor Center", "class_students", "class_student_id, class_id, child_profile_id, student_name, student_phone, student_email, status, enrolled_at, notes", "status: ENROLLED, DROPPED, COMPLETED"),
    ("UC-20", "Class Life-cycle", "Export Student via Excel", "Tutor Center", "class_students", "class_student_id, class_id, student_name, student_phone, student_email, status, enrolled_at", "status: ENROLLED, DROPPED, COMPLETED"),
    ("UC-21", "Class Life-cycle", "Monitor Active Class", "Platform Admin, Tutor Center", "tutoring_classes", "class_id, status, creator_id, subject_id, start_date, end_date", "status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("", "", "", "", "class_assignments", "assignment_id, application_id, status, assigned_date", "status: ACTIVE, TERMINATED"),
    ("UC-22", "Class Life-cycle", "Select Tutor for Class", "Client", "tutor_applications", "application_id, class_id, tutor_id, status, reviewed_at", "status: SUBMITTED, UNDER_REVIEW, ACCEPTED, REJECTED, WITHDRAWN"),
    ("", "", "", "", "class_assignments", "assignment_id, application_id, status", "status: ACTIVE, TERMINATED"),
    ("UC-23", "Class Life-cycle", "Manage Class Pipeline", "Tutor Center", "tutoring_classes", "class_id, status, title, subject_id, created_at", "status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("", "", "", "", "tutor_applications", "application_id, class_id, tutor_id, status", "status: SUBMITTED, UNDER_REVIEW, ACCEPTED, REJECTED, WITHDRAWN"),
    ("UC-24", "Class Life-cycle", "Manage Tutors", "Tutor Center", "center_tutor_memberships", "membership_id, center_id, tutor_id, recruitment_app_id, joined_at, status", "status: ACTIVE, INACTIVE, TERMINATED"),
    ("UC-25", "Class Life-cycle", "Manage Teaching Schedule", "Tutor Center", "schedule_slots", "slot_id, class_id, day_of_week, start_time, end_time", ""),
    ("", "", "", "", "lessons", "lesson_id, class_id, slot_id, sequence_no, attendance_status", "attendance_status: PENDING, COMPLETED, ABSENT, DISPUTED"),
    ("UC-26", "Class Life-cycle", "Request Tutor Replacement", "Client", "tutor_replacement_requests", "replacement_id, class_id, assignment_id, requested_by, preferred_tutor_id, reason, status, created_at, resolved_at", "status: PENDING, APPROVED, REJECTED, COMPLETED"),
    ("UC-27", "Class Life-cycle", "Manage Favorite Tutors", "Client, Tutor Center", "favorite_tutors", "favorite_id, user_id, tutor_id, created_at", ""),
    ("UC-28", "Class Life-cycle", "Request Early Class Termination", "Client, Tutor", "class_termination_requests", "termination_id, assignment_id, requested_by, reason, effective_date, status, created_at, processed_at", "status: PENDING, APPROVED, REJECTED, COMPLETED"),
    ("UC-29", "Class Life-cycle", "Report Class Disruption", "Client, Tutor, Tutor Center", "reports", "report_id, reporter_id, target_type, target_id, category, description, evidence_urls, status, created_at, updated_at", "target_type: USER, CLASS, REVIEW; category: FRAUD, ABUSE, SPAM; status: PENDING, RESOLVED"),
    ("UC-30", "Class Life-cycle", "Resolve Class Issues", "Tutor Center, Platform Admin", "support_tickets", "ticket_id, user_id, assigned_admin_id, target_class_id, category, subject, description, evidence_urls, priority, status, resolved_at, closed_at, created_at, updated_at", "category: DISPUTE, SYSTEM_ERROR, REPORT_USER, BUG_REPORT, INQUIRY; priority: LOW, MEDIUM, HIGH, URGENT; status: OPEN, IN_PROGRESS, IN_REVIEW, RESOLVED, CLOSED"),
    ("", "", "", "", "disputes", "dispute_id, report_id, escrow_id, resolution, status, created_at, updated_at", "status: OPEN, UNDER_INVESTIGATION, RESOLVED, WAITING"),
    # Search
    ("UC-31", "Advanced Search Engine", "Search Tutors", "Client, Tutor Center", "tutors", "tutor_id, user_id, full_name, gender, location_id, hourly_rate, rating_avg, verification_status, experience_years", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED; gender: MALE, FEMALE, OTHER"),
    ("", "", "", "", "tutor_subjects", "tutor_subject_id, tutor_id, subject_id, proficiency_level", ""),
    ("", "", "", "", "subjects", "subject_id, subject_name, description", ""),
    ("", "", "", "", "locations", "location_id, address_line, district_name, province_id, latitude, longitude", ""),
    ("", "", "", "", "matching_preferences", "preference_id, user_id, preferred_gender, min_rating, max_hourly_rate, preferred_location", ""),
    ("UC-32", "Advanced Search Engine", "Search Available Classes", "Tutor", "tutoring_classes", "class_id, subject_id, grade_id, location_id, lesson_mode, status, budget, tuition_fee, start_date", "lesson_mode: ONLINE, OFFLINE, HYBRID; status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("UC-33", "Advanced Search Engine", "View Detail Tutor", "Client, Tutor Center", "tutors", "tutor_id, user_id, full_name, gender, bio, avatar, experience_years, hourly_rate, rating_avg, verification_status, location_id", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("", "", "", "", "tutor_subjects", "tutor_subject_id, tutor_id, subject_id", ""),
    ("", "", "", "", "qualifications", "qualification_id, tutor_id, title, issuer, issue_date", ""),
    ("", "", "", "", "tutor_educations", "education_id, tutor_id, institution, degree", ""),
    ("", "", "", "", "tutor_certificates", "certificate_id, tutor_id, name, issuer", ""),
    ("UC-34", "Advanced Search Engine", "View Class Detail", "Tutor", "tutoring_classes", "class_id, creator_id, subject_id, grade_id, location_id, title, description, lesson_mode, number_of_sessions, tuition_fee, budget, status, start_date, end_date", "lesson_mode: ONLINE, OFFLINE, HYBRID; status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    # Messaging
    ("UC-35", "Real-Time Messaging & Notification", "Configure Notification Templates", "Platform Admin", "notification_templates", "template_id, code, title_template, content_template, channel", ""),
    ("UC-36", "Real-Time Messaging & Notification", "Chat with Users", "Client, Tutor, Tutor Center", "conversations", "conversation_id, context_type, context_id, type, status, last_message_at, created_at", "status: ACTIVE, ARCHIVED"),
    ("", "", "", "", "conversation_participants", "participant_id, conversation_id, user_id", ""),
    ("", "", "", "", "messages", "message_id, conversation_id, sender_id, message_type, content, is_edited, edited_at, is_deleted, sent_at", "message_type: TEXT, IMAGE, FILE"),
    ("", "", "", "", "message_attachments", "attachment_id, message_id, file_id, created_at", ""),
    ("UC-37", "Real-Time Messaging & Notification", "View Notifications", "Client, Tutor, Tutor Center, Platform Admin", "notifications", "notification_id, user_id, template_id, type, title, content, reference_type, reference_id, status, is_read, read_at, created_at", "type: PAYMENT, APPLICATION, SYSTEM, CLASS, VERIFICATION, CHAT; status: PENDING, SENT, READ, FAILED"),
    ("", "", "", "", "notification_preferences", "preference_id, user_id, email_enabled, sms_enabled, push_enabled", ""),
    ("", "", "", "", "notification_queues", "queue_id, notification_id, channel, status, scheduled_at, sent_at", ""),
    # Escrow Wallet
    ("UC-38", "Smart Escrow Wallet", "Manage Payment Method", "Client, Tutor, Tutor Center", "payment_methods", "payment_method_id, wallet_id, type, account_no, bank_name, status", "status: ACTIVE"),
    ("", "", "", "", "wallets", "wallet_id, available_balance, frozen_balance, status, created_at, updated_at", "status: ACTIVE, SUSPENDED, CLOSED"),
    ("UC-39", "Smart Escrow Wallet", "Manage Wallet", "Client, Tutor, Tutor Center", "wallets", "wallet_id, available_balance, frozen_balance, status, created_at, updated_at", "status: ACTIVE, SUSPENDED, CLOSED"),
    ("", "", "", "", "financial_journals", "journal_id, wallet_id, reference_type, reference_id, entry_type, amount, balance_before, balance_after, created_at", "entry_type: CREDIT, DEBIT"),
    ("UC-40", "Smart Escrow Wallet", "Manage Escrow Payment", "Client, Tutor, Tutor Center, Platform Admin", "escrow_transactions", "escrow_id, payment_id, assignment_id, amount, status, deposited_at, released_at, created_at, updated_at", "status: PENDING, FUNDED, RELEASED, REFUNDED, ON_HOLD, DISPUTED"),
    ("", "", "", "", "payment_transactions", "transaction_id, wallet_id, external_transaction_id, type, status, amount, description, reference_code, processed_at, created_at", "type: DEPOSIT, WITHDRAWAL, REFUND, ESCROW_DEPOSIT, ESCROW_RELEASE; status: PENDING, SUCCESS, FAILED, CANCELLED"),
    ("UC-41", "Smart Escrow Wallet", "Monitor Financial Report", "Tutor Center, Platform Admin", "financial_journals", "journal_id, wallet_id, reference_type, reference_id, entry_type, amount, balance_before, balance_after, created_at", "entry_type: CREDIT, DEBIT"),
    ("", "", "", "", "payment_transactions", "transaction_id, wallet_id, type, status, amount, created_at", "type: DEPOSIT, WITHDRAWAL, REFUND, ESCROW_DEPOSIT, ESCROW_RELEASE; status: PENDING, SUCCESS, FAILED, CANCELLED"),
    ("", "", "", "", "escrow_transactions", "escrow_id, assignment_id, amount, status, deposited_at, released_at", "status: PENDING, FUNDED, RELEASED, REFUNDED, ON_HOLD, DISPUTED"),
    ("UC-42", "Smart Escrow Wallet", "Handle Stale Escrow & Auto-Refund", "System", "escrow_transactions", "escrow_id, status, deposited_at, amount, assignment_id", "status: PENDING, FUNDED, ON_HOLD, REFUNDED"),
    ("", "", "", "", "refund_requests", "refund_id, escrow_id, requested_by, reason, amount, status, requested_at, processed_at", "status: PENDING, APPROVED, REJECTED, COMPLETED"),
    ("UC-43", "Smart Escrow Wallet", "Export Financial Statements", "Tutor Center, Platform Admin", "financial_journals", "journal_id, wallet_id, entry_type, amount, balance_before, balance_after, created_at", "entry_type: CREDIT, DEBIT"),
    ("UC-44", "Smart Escrow Wallet", "Manage E-Contract", "Client, Tutor, Tutor Center", "contracts", "contract_id, contract_no, assignment_id, template_id, contract_file_url, terms_summary, status, signed_at, created_at, updated_at", "status: DRAFT, SIGNED, ACTIVE, COMPLETED, TERMINATED"),
    ("", "", "", "", "contract_signatures", "signature_id, contract_id, signer_id, signed_at, signature_data", ""),
    ("UC-45", "Smart Escrow Wallet", "Manage E-Contract Templates", "Tutor Center, Platform Admin", "contract_templates", "template_id, name, content, created_by, center_id, is_default, status, created_at, updated_at", "status: DRAFT, ACTIVE, ARCHIVED"),
    ("UC-46", "Smart Escrow Wallet", "Configure Platform Fees", "Platform Admin", "system_parameters", "parameter_id, param_key, param_value, description", ""),
    ("UC-47", "Smart Escrow Wallet", "View Earning Report", "Tutor", "payment_release_requests", "request_id, assignment_id, amount, status, requested_at, processed_at", "status: PENDING, APPROVED, REJECTED, PAID"),
    ("", "", "", "", "financial_journals", "journal_id, wallet_id, entry_type, amount, created_at", "entry_type: CREDIT, DEBIT"),
    # Dispute
    ("UC-48", "Dispute & Resolution Management", "Create Dispute Ticket", "Client, Tutor, Tutor Center", "reports", "report_id, reporter_id, target_type, target_id, category, description, evidence_urls, status", "target_type: USER, CLASS, REVIEW; category: FRAUD, ABUSE, SPAM; status: PENDING, RESOLVED"),
    ("", "", "", "", "disputes", "dispute_id, report_id, escrow_id, resolution, status, created_at, updated_at", "status: OPEN, UNDER_INVESTIGATION, RESOLVED, WAITING"),
    ("UC-49", "Dispute & Resolution Management", "Manage Dispute Ticket", "Platform Admin", "disputes", "dispute_id, report_id, escrow_id, resolution, status, created_at, updated_at", "status: OPEN, UNDER_INVESTIGATION, RESOLVED, WAITING"),
    ("UC-50", "Dispute & Resolution Management", "Request Refund", "Client, Tutor, Tutor Center", "refund_requests", "refund_id, escrow_id, requested_by, reason, amount, status, requested_at, processed_at", "status: PENDING, APPROVED, REJECTED, COMPLETED"),
    ("UC-51", "Dispute & Resolution Management", "Manage Refund Requests", "Platform Admin", "refund_requests", "refund_id, escrow_id, requested_by, reason, amount, status, requested_at, processed_at", "status: PENDING, APPROVED, REJECTED, COMPLETED"),
    ("UC-52", "Dispute & Resolution Management", "Report User or Class", "Client, Tutor, Tutor Center", "reports", "report_id, reporter_id, target_type, target_id, category, description, evidence_urls, status, created_at, updated_at", "target_type: USER, CLASS, REVIEW; category: FRAUD, ABUSE, SPAM; status: PENDING, RESOLVED"),
    # Review
    ("UC-53", "Review & Reputation Analytics Engine", "View Rating & Review", "Client, Tutor, Tutor Center, Platform Admin", "reviews", "review_id, assignment_id, class_id, reviewer_id, reviewee_id, review_type, rating, comment, status, created_at", "review_type: CLIENT_TO_TUTOR, TUTOR_TO_CLIENT, CENTER_TO_TUTOR; status: VISIBLE, HIDDEN, MODERATED"),
    ("", "", "", "", "reputation_histories", "history_id, tutor_id, old_score, new_score, trigger_type, reason, created_at", ""),
    ("", "", "", "", "tutors", "tutor_id, rating_avg", ""),
    ("UC-54", "Review & Reputation Analytics Engine", "Submit Rating & Review", "Client, Tutor, Tutor Center", "reviews", "review_id, assignment_id, reviewer_id, reviewee_id, review_type, rating, comment, status, created_at", "review_type: CLIENT_TO_TUTOR, TUTOR_TO_CLIENT, CENTER_TO_TUTOR; status: VISIBLE, HIDDEN, MODERATED"),
    ("UC-55", "Review & Reputation Analytics Engine", "Manage Review & Feedback", "Platform Admin", "reviews", "review_id, assignment_id, status, comment, rating", "status: VISIBLE, HIDDEN, MODERATED"),
    # Operations
    ("UC-56", "Operations Dashboard & Monitoring", "View Dashboard", "Platform Admin", "users", "user_id, status, created_at", "status: ACTIVE, SUSPENDED, BANNED"),
    ("", "", "", "", "tutoring_classes", "class_id, status", "status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("", "", "", "", "escrow_transactions", "escrow_id, status, amount", "status: PENDING, FUNDED, RELEASED, REFUNDED, ON_HOLD, DISPUTED"),
    ("", "", "", "", "support_tickets", "ticket_id, status, priority", "status: OPEN, IN_PROGRESS, IN_REVIEW, RESOLVED, CLOSED"),
    ("UC-57", "Operations Dashboard & Monitoring", "Manage System Categories", "Platform Admin", "categories", "category_id, parent_id, name, type, description, is_active, sort_order, status, created_at, updated_at", "type: SUBJECT, EDUCATION_LEVEL, LOCATION, SYSTEM_CONFIG"),
    ("UC-58", "Operations Dashboard & Monitoring", "Manage Escrow Transactions", "Platform Admin", "escrow_transactions", "escrow_id, payment_id, assignment_id, amount, status, deposited_at, released_at, created_at, updated_at", "status: PENDING, FUNDED, RELEASED, REFUNDED, ON_HOLD, DISPUTED"),
    ("", "", "", "", "payment_histories", "payment_history_id, payment_id, old_status, new_status, changed_by_user_id, changed_at", ""),
    ("UC-59", "Operations Dashboard & Monitoring", "Detect Platform Circumvention", "Platform Admin", "user_penalties", "penalty_id, user_id, penalty_type, reason, status, starts_at, expires_at", "penalty_type: WARNING, FEATURE_RESTRICTION, TEMPORARY_BAN, PERMANENT_BAN"),
    ("", "", "", "", "audit_logs", "audit_id, actor_id, action, entity_type, entity_id, created_at", ""),
    ("", "", "", "", "reports", "report_id, target_type, category, status", "category: FRAUD, ABUSE, SPAM"),
    ("UC-60", "Operations Dashboard & Monitoring", "Enforce Platform Penalties", "Platform Admin", "user_penalties", "penalty_id, user_id, issued_by, penalty_type, reason, evidence_urls, restriction_details, starts_at, expires_at, status, revoked_at, revoked_reason", "penalty_type: WARNING, FEATURE_RESTRICTION, TEMPORARY_BAN, PERMANENT_BAN; status: ACTIVE, EXPIRED, REVOKED"),
    ("", "", "", "", "users", "user_id, status", "status: ACTIVE, SUSPENDED, BANNED"),
    ("UC-61", "Operations Dashboard & Monitoring", "Monitor Audit Logs", "Platform Admin", "audit_logs", "audit_id, actor_id, action, entity_type, entity_id, old_value, new_value, ip_address, user_agent, created_at", ""),
    # AI & Support
    ("UC-62", "Intelligent Dynamic Matchmaking", "AI-Recommended Tutors", "Client, Tutor Center", "recommendation_logs", "recommendation_id, user_id, tutor_id, class_id, score, algorithm_version, reason, generated_at", ""),
    ("", "", "", "", "matching_preferences", "preference_id, user_id, preferred_gender, min_rating, max_hourly_rate, preferred_location", ""),
    ("", "", "", "", "tutors", "tutor_id, rating_avg, verification_status, hourly_rate", "verification_status: UNDER_VERIFY, VERIFIED, REJECTED"),
    ("UC-63", "Intelligent Dynamic Matchmaking", "AI-Recommended Jobs", "Tutor", "recommendation_logs", "recommendation_id, user_id, tutor_id, recruitment_id, score, algorithm_version, reason, generated_at", ""),
    ("", "", "", "", "recruitment_posts", "recruitment_id, center_id, title, subject_id, location_id, status, published_at", "status: DRAFT, ACTIVE, CLOSED"),
    ("UC-64", "Intelligent Dynamic Matchmaking", "AI-Recommended Classes", "Tutor", "recommendation_logs", "recommendation_id, user_id, tutor_id, class_id, score, algorithm_version, reason, generated_at", ""),
    ("", "", "", "", "tutoring_classes", "class_id, subject_id, grade_id, location_id, status, budget", "status: DRAFT, OPEN, MATCHED, IN_PROGRESS, COMPLETED, CANCELLED, DISPUTED"),
    ("UC-65", "Smart Chatbot Advisor", "Get Support", "Client, Tutor, Tutor Center", "faq_entries", "faq_id, question, answer, category, sort_order, is_published, created_at, updated_at", ""),
    ("", "", "", "", "support_tickets", "ticket_id, user_id, category, subject, description, status", "category: DISPUTE, SYSTEM_ERROR, REPORT_USER, BUG_REPORT, INQUIRY; status: OPEN, IN_PROGRESS, IN_REVIEW, RESOLVED, CLOSED"),
    ("UC-66", "Smart Chatbot Advisor", "Manage Support Requests", "Platform Admin", "support_tickets", "ticket_id, user_id, assigned_admin_id, target_class_id, category, subject, description, evidence_urls, priority, status, resolved_at, closed_at, created_at, updated_at", "category: DISPUTE, SYSTEM_ERROR, REPORT_USER, BUG_REPORT, INQUIRY; priority: LOW, MEDIUM, HIGH, URGENT; status: OPEN, IN_PROGRESS, IN_REVIEW, RESOLVED, CLOSED"),
    ("UC-67", "Smart Chatbot Advisor", "Manage FAQ Knowledge Base", "Platform Admin", "faq_entries", "faq_id, question, answer, category, sort_order, is_published, created_at, updated_at", ""),
    # Recruitment
    ("UC-68", "Tutor Recruitment", "Manage Tutor Recruitment", "Tutor Center", "recruitment_posts", "recruitment_id, center_id, title, description, requirements, benefits, required_experience, subject_id, location_id, max_positions, status, published_at, closed_at, created_at, updated_at", "status: DRAFT, ACTIVE, CLOSED"),
    ("UC-69", "Tutor Recruitment", "Apply Recruitment", "Tutor", "recruitment_applications", "recruitment_app_id, recruitment_id, tutor_id, cover_letter, resume_url, status, interview_date, interview_notes, applied_at, reviewed_at, updated_at", "status: APPLIED, SCREENING, INTERVIEW, PASSED, HIRED, REJECTED, WITHDRAWN"),
    ("UC-70", "Tutor Recruitment", "Search Recruitment Posts", "Tutor", "recruitment_posts", "recruitment_id, center_id, title, subject_id, location_id, required_experience, status, published_at", "status: DRAFT, ACTIVE, CLOSED"),
    ("UC-71", "Tutor Recruitment", "Manage Recruitment Applications", "Tutor Center", "recruitment_applications", "recruitment_app_id, recruitment_id, tutor_id, cover_letter, resume_url, status, interview_date, interview_notes, applied_at, reviewed_at, updated_at", "status: APPLIED, SCREENING, INTERVIEW, PASSED, HIRED, REJECTED, WITHDRAWN"),
    ("", "", "", "", "center_tutor_memberships", "membership_id, center_id, tutor_id, recruitment_app_id, joined_at, status", "status: ACTIVE, INACTIVE, TERMINATED"),
]

COLUMN_WIDTHS = [10, 32, 36, 36, 28, 80, 60]


def main() -> None:
    out = Path(__file__).resolve().parent / "UC_Entity_Mapping_71UC.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "UC Entity Mapping"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(ROWS) + 1}"

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(out)
    print(f"Created: {out} ({len(ROWS)} rows)")


if __name__ == "__main__":
    main()
